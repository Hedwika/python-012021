import openpyxl
import pandas
import pandas as pd

# Ulož tabulku s cekovými počty vykázaných hodin, kterou jsi vytvořila v příkladu 27 jako Excel soubor. Pokud jsi tento příklad nezpracovala, ulož libovolnou jinou tabulku jako Excel sešit. Výsledný sešit si otevři a zkontroluj. K uložení použij funkci to_excel, se kterou pracuj stejně, jako jsme na lekci pracovali s funkci to_csv.
vykazy = pandas.read_csv('vykazy.csv')
hodiny = vykazy.groupby('project').sum("hours")
hodiny = hodiny.drop(columns="emloyee_id")
df = pd.DataFrame(hodiny)
df.to_excel('hodiny.xlsx')
hodiny_xls = pandas.read_excel('hodiny.xlsx')
print(hodiny_xls)

# Rozšířené zadání
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]

sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
modra_barva = PatternFill("solid", fgColor="009999FF")
staroruzova_barva = PatternFill("solid", fgColor="00993366")
zluta_barva = PatternFill("solid", fgColor="00FFFF00")
zelena_barva = PatternFill("solid", fgColor="00339966")

radek = 1
i = 0
for den in rozvrh_hodin:
    sloupec = 1
    j = 0
    for predmet in den:
        bunka = ws1.cell(radek, sloupec)
        bunka.value = rozvrh_hodin[i][j]
        if bunka.value == "Oběd":
            bunka.fill = sediva_barva
        if bunka.value == "Anglický jazyk":
            bunka.fill = modra_barva
        if bunka.value == "Matematika":
            bunka.fill = staroruzova_barva
        if bunka.value == "Výtvarná výchova":
            bunka.fill = zluta_barva
        if bunka.value == "Přírodopis":
            bunka.fill = zelena_barva
        j += 1
        sloupec += 1
    radek += 1
    i += 1

# den = 0
# predmet = 0
# radek = 1
# while den < len(rozvrh_hodin):
#     sloupec = 1
#     while predmet < len(rozvrh_hodin[den]):
#         bunka = ws1.cell(radek, sloupec)
#         bunka.value = rozvrh_hodin[den][predmet]
#         if bunka.value == "Oběd":
#             bunka.fill = sediva_barva
#         if bunka.value == "Anglický jazyk":
#             bunka.fill = modra_barva
#         if bunka.value == "Matematika":
#             bunka.fill = staroruzova_barva
#         if bunka.value == "Výtvarná výchova":
#             bunka.fill = zluta_barva
#         if bunka.value == "Přírodopis":
#             bunka.fill = zelena_barva
#         predmet += 1
#         sloupec += 1
#         if predmet == len(rozvrh_hodin[den]):
#             den += 1
#             predmet = 0
#             radek += 1
#             sloupec = 1
#             if den == len(rozvrh_hodin):
#                 break

wb.save(filename="rozvrh_hodin.xlsx")